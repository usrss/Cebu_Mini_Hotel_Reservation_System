"""
reports/services.py

Enhanced ReportService for the custom report generation feature.

Extends the existing staff/services.py with:
  - Metric selection  — only return requested metrics
  - Flexible grouping — day | week | month | room_type | status
  - Role-based scoping — admin = system-wide, manager = limited
  - Result caching     — django.core.cache (default backend)
  - PDF export         — via ReportLab
  - Excel export       — via openpyxl

All methods return a standard envelope:
  {
    "summary": { ... selected metrics ... },
    "rows":    [ ... grouped rows ... ],
    "meta": {
      "report_type": ...,
      "start_date":  ...,
      "end_date":    ...,
      "group_by":    ...,
      "metrics":     [...],
      "generated_at": ...,
      "cached":      bool,
    }
  }
"""

import hashlib
import io
import json
import logging
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.cache   import cache
from django.db           import models
from django.db.models    import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils        import timezone

from bookings.models import Booking, BookingStatus
from bookings.models import PaymentStatus as BookingPaymentStatus
from rooms.models    import Room

logger = logging.getLogger(__name__)
User   = get_user_model()

CACHE_TTL = 300  # 5 minutes


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _trunc_fn(group_by: str):
    """Return the ORM truncation function for a group_by value."""
    return {
        "day":   TruncDate,
        "week":  TruncWeek,
        "month": TruncMonth,
    }.get(group_by, TruncDate)


def _cache_key(report_type: str, config: dict) -> str:
    payload = json.dumps({"t": report_type, "c": config}, sort_keys=True)
    return f"report:{hashlib.md5(payload.encode()).hexdigest()}"


def _filter_summary(summary: dict, metrics: list) -> dict:
    """Return only the requested metrics from the summary dict."""
    if not metrics:
        return summary
    return {k: v for k, v in summary.items() if k in metrics}


def _resolve_dates(config: dict):
    """Resolve start_date and end_date from config."""
    today  = timezone.now().date()
    period = config.get("period", "monthly")

    if period == "custom":
        start = date.fromisoformat(config["start_date"])
        end   = date.fromisoformat(config["end_date"])
        return start, end

    mapping = {
        "daily":   (today,                         today),
        "weekly":  (today - timedelta(days=6),     today),
        "monthly": (today.replace(day=1),          today),
        "yearly":  (today.replace(month=1, day=1), today),
    }
    return mapping.get(period, mapping["monthly"])


# ─── EnhancedReportService ────────────────────────────────────────────────────

class EnhancedReportService:

    # ── Public dispatcher ──────────────────────────────────────────────────────

    @classmethod
    def run(cls, report_type: str, config: dict, user=None) -> dict:
        """
        Main entry point.
        Checks cache first, calls the appropriate generator, caches result.
        """
        key    = _cache_key(report_type, config)
        cached = cache.get(key)
        if cached:
            cached["meta"]["cached"] = True
            return cached

        start, end = _resolve_dates(config)
        metrics    = config.get("metrics", [])
        group_by   = config.get("group_by", "day")
        filters    = config.get("filters", {})

        generators = {
            "bookings":  cls._bookings,
            "revenue":   cls._revenue,
            "occupancy": cls._occupancy,
            "guests":    cls._guests,
            "staff":     cls._staff,
            "food":      cls._food,
            "payments":  cls._payments,
        }

        fn = generators.get(report_type)
        if not fn:
            raise ValueError(f"Unknown report_type: {report_type}")

        data = fn(start, end, metrics=metrics, group_by=group_by, filters=filters, user=user)
        data["meta"] = {
            "report_type":  report_type,
            "start_date":   str(start),
            "end_date":     str(end),
            "group_by":     group_by,
            "metrics":      metrics,
            "generated_at": timezone.now().isoformat(),
            "cached":       False,
        }

        cache.set(key, data, CACHE_TTL)
        return data

    # ── Bookings ───────────────────────────────────────────────────────────────

    @staticmethod
    def _bookings(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        qs = Booking.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        )

        # Apply optional filters
        if filters.get("room_type"):
            qs = qs.filter(room__room_type=filters["room_type"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])

        summary_full = {
            "total":         qs.count(),
            "pending_payment": qs.filter(status=BookingStatus.PENDING_PAYMENT).count(),
            "confirmed":     qs.filter(status=BookingStatus.CONFIRMED).count(),
            "checked_in":    qs.filter(status=BookingStatus.CHECKED_IN).count(),
            "checked_out":   qs.filter(status=BookingStatus.CHECKED_OUT).count(),
            "cancelled":     qs.filter(status=BookingStatus.CANCELLED).count(),
            "expired":       qs.filter(status=BookingStatus.EXPIRED).count(),
            "no_show":       qs.filter(status=BookingStatus.NO_SHOW).count(),
            "total_revenue": float(
                qs.filter(payment_status=BookingPaymentStatus.PAID)
                  .aggregate(t=Sum("total_price"))["t"] or 0
            ),
        }

        # Grouping
        if group_by in ("day", "week", "month"):
            trunc = _trunc_fn(group_by)
            rows_qs = (
                qs.annotate(period=trunc("created_at"))
                  .values("period")
                  .annotate(
                      count=Count("id"),
                      paid_revenue=Sum(
                          "total_price",
                          filter=Q(payment_status=BookingPaymentStatus.PAID),
                      ),
                  )
                  .order_by("period")
            )
            rows = [
                {
                    "period":       str(r["period"]),
                    "bookings":     r["count"],
                    "paid_revenue": float(r["paid_revenue"] or 0),
                }
                for r in rows_qs
            ]
        elif group_by == "status":
            rows = [
                {"status": s, "count": qs.filter(status=s).count()}
                for s in BookingStatus.values
            ]
        elif group_by == "room_type":
            from rooms.models import RoomType
            rows = [
                {
                    "room_type":  rt,
                    "label":      label,
                    "count":      qs.filter(room__room_type=rt).count(),
                }
                for rt, label in RoomType.choices
            ]
        else:
            rows = []

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }

    # ── Revenue ────────────────────────────────────────────────────────────────

    @staticmethod
    def _revenue(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        paid_qs = Booking.objects.filter(
            payment_status=BookingPaymentStatus.PAID,
            confirmed_at__date__gte=start,
            confirmed_at__date__lte=end,
        )

        if filters.get("room_type"):
            paid_qs = paid_qs.filter(room__room_type=filters["room_type"])

        totals = paid_qs.aggregate(
            total_revenue    = Sum("total_price"),
            total_tax        = Sum("tax"),
            total_service_fee= Sum("service_fee"),
            avg_booking_value= Avg("total_price"),
            count            = Count("id"),
        )

        refund_total = (
            Booking.objects.filter(
                refund_status="completed",
                cancelled_at__date__gte=start,
                cancelled_at__date__lte=end,
            ).aggregate(t=Sum("refund_amount"))["t"] or Decimal("0")
        )

        summary_full = {
            "total_revenue":     float(totals["total_revenue"] or 0),
            "total_tax":         float(totals["total_tax"] or 0),
            "total_service_fee": float(totals["total_service_fee"] or 0),
            "net_revenue":       float((totals["total_revenue"] or 0) - refund_total),
            "avg_booking_value": float(totals["avg_booking_value"] or 0),
            "paid_bookings":     totals["count"] or 0,
            "total_refunds":     float(refund_total),
        }

        # Grouping
        if group_by in ("day", "week", "month"):
            trunc = _trunc_fn(group_by)
            rows_qs = (
                paid_qs.annotate(period=trunc("confirmed_at"))
                       .values("period")
                       .annotate(revenue=Sum("total_price"), bookings=Count("id"))
                       .order_by("period")
            )
            rows = [
                {
                    "period":   str(r["period"]),
                    "revenue":  float(r["revenue"] or 0),
                    "bookings": r["bookings"],
                }
                for r in rows_qs
            ]
        elif group_by == "room_type":
            from rooms.models import RoomType
            rows = []
            for rt, label in RoomType.choices:
                sub = paid_qs.filter(room__room_type=rt)
                rev = sub.aggregate(t=Sum("total_price"))["t"] or 0
                rows.append({
                    "room_type": rt,
                    "label":     label,
                    "revenue":   float(rev),
                    "bookings":  sub.count(),
                })
        else:
            rows = []

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }

    # ── Occupancy ──────────────────────────────────────────────────────────────

    @staticmethod
    def _occupancy(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        rooms_qs = Room.objects.filter(is_active=True)
        if filters.get("room_type"):
            rooms_qs = rooms_qs.filter(room_type=filters["room_type"])

        total_rooms       = rooms_qs.count()
        total_days        = (end - start).days + 1
        total_room_nights = total_rooms * total_days

        overlapping = Booking.objects.filter(
            status__in=[
                BookingStatus.CONFIRMED,
                BookingStatus.CHECKED_IN,
                BookingStatus.CHECKED_OUT,
            ],
            check_in__lte=end,
            check_out__gte=start,
        )
        if filters.get("room_type"):
            overlapping = overlapping.filter(room__room_type=filters["room_type"])

        occupied_nights = 0
        for b in overlapping:
            actual_start    = max(b.check_in, start)
            actual_end      = min(b.check_out, end)
            occupied_nights += (actual_end - actual_start).days

        occupancy_rate = (
            round(occupied_nights / total_room_nights * 100, 2)
            if total_room_nights > 0 else 0.0
        )

        summary_full = {
            "total_rooms":       total_rooms,
            "total_days":        total_days,
            "total_room_nights": total_room_nights,
            "occupied_nights":   occupied_nights,
            "occupancy_rate":    occupancy_rate,
            # Period-average daily occupancy rate (percent).
            "avg_occupancy_rate": occupancy_rate,
        }

        # By room type
        from rooms.models import RoomType
        rows = []
        for rt_key, rt_label in RoomType.choices:
            type_rooms = rooms_qs.filter(room_type=rt_key)
            count = type_rooms.count()
            if count == 0:
                continue
            booked      = overlapping.filter(room__room_type=rt_key)
            type_nights = 0
            for b in booked:
                actual_start = max(b.check_in, start)
                actual_end   = min(b.check_out, end)
                type_nights += (actual_end - actual_start).days
            max_nights = count * total_days
            rows.append({
                "room_type":       rt_key,
                "room_type_label": rt_label,
                "room_count":      count,
                "occupied_nights": type_nights,
                "max_nights":      max_nights,
                "occupancy_rate":  round(type_nights / max_nights * 100, 2) if max_nights else 0,
            })

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }

    # ── Guests ─────────────────────────────────────────────────────────────────

    @staticmethod
    def _guests(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        new_users = User.objects.filter(
            date_joined__date__gte=start,
            date_joined__date__lte=end,
            is_staff=False,
        ).count()

        repeat_guests = (
            Booking.objects.filter(payment_status=BookingPaymentStatus.PAID)
            .values("user")
            .annotate(c=Count("id"))
            .filter(c__gt=1)
            .count()
        )

        period_qs = Booking.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
        walk_ins   = period_qs.filter(user__isnull=True).count()
        registered = period_qs.filter(user__isnull=False).count()

        top_guests = (
            period_qs.filter(user__isnull=False)
            .values("user__email", "full_name")
            .annotate(bookings=Count("id"), spent=Sum("total_price"))
            .order_by("-bookings")[:10]
        )

        summary_full = {
            "new_registrations":   new_users,
            "repeat_guests":       repeat_guests,
            "walk_in_bookings":    walk_ins,
            "registered_bookings": registered,
        }

        rows = [
            {
                "email":    g["user__email"],
                "name":     g["full_name"],
                "bookings": g["bookings"],
                "spent":    float(g["spent"] or 0),
            }
            for g in top_guests
        ]

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }

    # ── Staff Performance ──────────────────────────────────────────────────────

    @staticmethod
    def _staff(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        """
        Returns per-staff breakdown rows[] that StaffResultPanel can consume.

        Each row contains:
          staff_id, name, email, role,
          check_ins_handled, check_outs_handled, bookings_created,
          cancellations_processed, avg_check_in_time,   ← front_desk
          rooms_cleaned, avg_cleaning_time,
          rooms_cleaned_per_shift, delayed_cleanings,   ← housekeeping
          orders_completed, avg_preparation_time,
          pending_orders, cancelled_orders,             ← kitchen_staff

        Role detection on the frontend (StaffResultPanel.detectRole) reads the
        explicit `role` field, so we must populate it — the old code never did,
        which caused every row to fall through all heuristics and be silently
        dropped, leaving the leaderboard empty.

        Business rules:
          - All active, in-scope StaffProfile rows appear even if inactive
            in the period (values default to 0, not NULL).
          - front_desk / receptionist  → check-in metrics from StaffActivityLog
          - housekeeping               → CleaningTask completions
          - maintenance                → MaintenanceTask completions
          - kitchen_staff              → FoodOrder completions (if food app present)
          - admin / manager            → excluded from per-staff rows
                                         (they appear in scoped_profiles but are
                                          filtered below since the panel has no
                                          role section for them)
          - Managers cannot see admin/manager profile data (role-scope guard).
        """
        from staff.models import (
            StaffActivityLog, CleaningTask, MaintenanceTask,
            StaffProfile, StaffRole,
        )

        caller_role = (
            getattr(getattr(user, "staff_profile", None), "effective_role", None)
            if user else None
        )

        # ── 1. Resolve in-scope profiles ──────────────────────────────────────
        profiles_qs = (
            StaffProfile.objects
            .select_related("user")
            .filter(is_active=True)
        )
        if caller_role == "manager":
            profiles_qs = profiles_qs.exclude(role__in=["admin", "manager"])
        if filters.get("role"):
            profiles_qs = profiles_qs.filter(role=filters["role"])
        if filters.get("staff_id"):
            profiles_qs = profiles_qs.filter(pk=filters["staff_id"])

        # Exclude admin/manager from the per-staff breakdown rows — there are no
        # role sections for them in StaffResultPanel.
        OPERATIONAL_ROLES = {
            StaffRole.FRONT_DESK,
            StaffRole.RECEPTIONIST,   # receptionist maps to front_desk panel
            StaffRole.HOUSEKEEPING,
            StaffRole.MAINTENANCE,
            StaffRole.KITCHEN_STAFF,
            StaffRole.SECURITY,
        }
        profiles_qs = profiles_qs.filter(role__in=OPERATIONAL_ROLES)

        # Build a dict keyed by profile.pk so we can merge activity counts in.
        # Seed every profile with zero counts so inactive staff still appear.
        perf: dict[int, dict] = {}
        for p in profiles_qs:
            full_name = p.user.get_full_name() or p.user.email

            # Map receptionist → front_desk so the frontend panel picks it up.
            panel_role = (
                "front_desk"
                if p.role == StaffRole.RECEPTIONIST
                else p.role
            )

            perf[p.pk] = {
                # Identity
                "staff_id":  p.employee_id or str(p.pk),
                "name":      full_name,
                "email":     p.user.email,
                "role":      panel_role,   # ← what StaffResultPanel.detectRole() reads
                # front_desk metrics
                "check_ins_handled":        0,
                "check_outs_handled":       0,
                "bookings_created":         0,
                "cancellations_processed":  0,
                "avg_check_in_time":        None,
                # housekeeping metrics
                "rooms_cleaned":            0,
                "avg_cleaning_time":        None,
                "rooms_cleaned_per_shift":  0,
                "delayed_cleanings":        0,
                # maintenance / kitchen metrics
                "orders_completed":         0,
                "avg_preparation_time":     None,
                "pending_orders":           0,
                "cancelled_orders":         0,
                # security metrics
                "incidents_logged":         0,
                "incidents_resolved":       0,
                "high_severity":            0,
                "avg_resolution_time":      None,
            }

        profile_ids = list(perf.keys())
        if not profile_ids:
            return {
                "summary": _filter_summary(
                    {"total_check_ins": 0, "total_cleaning_done": 0,
                     "total_maintenance_done": 0, "total_staff": 0},
                    metrics,
                ),
                "rows": [],
            }

        # ── 2. Front-desk metrics from StaffActivityLog ───────────────────────
        # Each action_type maps to a counter on the row.
        ACTION_MAP = {
            "check_in_guest":          "check_ins_handled",
            "check_out_guest":         "check_outs_handled",
            "create_booking":          "bookings_created",
            "cancel_booking":          "cancellations_processed",
            # Legacy aliases that may exist in older audit entries
            "guest_check_in":          "check_ins_handled",
            "guest_check_out":         "check_outs_handled",
        }
        activity_rows = (
            StaffActivityLog.objects.filter(
                action_type__in=list(ACTION_MAP.keys()),
                created_at__date__gte=start,
                created_at__date__lte=end,
                staff_id__in=profile_ids,
            )
            .values("staff_id", "action_type")
            .annotate(count=Count("id"))
        )
        for row in activity_rows:
            pid  = row["staff_id"]
            key  = ACTION_MAP.get(row["action_type"])
            if pid in perf and key:
                perf[pid][key] += row["count"]

        # ── 3. Housekeeping — CleaningTask completions ────────────────────────
        cleaning_rows = (
            CleaningTask.objects.filter(
                status="clean",
                completed_at__date__gte=start,
                completed_at__date__lte=end,
                assigned_to_id__in=profile_ids,
            )
            .values("assigned_to_id")
            .annotate(
                count=Count("id"),
                avg_mins=Avg(
                    F("completed_at") - F("started_at"),
                    output_field=models.DurationField(),
                ),
                delayed=Count(
                    "id",
                    filter=Q(completed_at__gt=F("cleaning_end_at"),
                             cleaning_end_at__isnull=False),
                ),
            )
        )
        for row in cleaning_rows:
            pid = row["assigned_to_id"]
            if pid not in perf:
                continue
            perf[pid]["rooms_cleaned"] = row["count"]
            if row["avg_mins"] is not None:
                # DurationField returns a timedelta
                total_secs = row["avg_mins"].total_seconds()
                perf[pid]["avg_cleaning_time"] = round(total_secs / 60, 1)
            perf[pid]["delayed_cleanings"] = row["delayed"]

        # rooms_cleaned_per_shift: divide by number of distinct shift days in period
        shift_days_in_period = max((end - start).days + 1, 1)
        for pid in perf:
            if perf[pid]["rooms_cleaned"] > 0:
                perf[pid]["rooms_cleaned_per_shift"] = round(
                    perf[pid]["rooms_cleaned"] / shift_days_in_period, 2
                )

        # ── 4. Maintenance — MaintenanceTask completions ──────────────────────
        maintenance_rows = (
            MaintenanceTask.objects.filter(
                status="completed",
                completed_at__date__gte=start,
                completed_at__date__lte=end,
                assigned_to_id__in=profile_ids,
            )
            .values("assigned_to_id")
            .annotate(count=Count("id"))
        )
        for row in maintenance_rows:
            pid = row["assigned_to_id"]
            if pid in perf:
                perf[pid]["orders_completed"] = row["count"]   # reuse orders_completed field
                # StaffResultPanel uses orders_completed for maintenance score too —
                # it reads ROLE_CONFIG[maintenance].scoreKeys = ['orders_completed']

        # ── 5. Kitchen staff — FoodOrder completions (optional) ───────────────
        try:
            from food.models import FoodOrder
            kitchen_ids = [
                pid for pid, d in perf.items()
                if d["role"] == "kitchen_staff"
            ]
            if kitchen_ids:
                # FoodOrder may link to staff via a FK; gracefully skip if not.
                food_qs_kwargs = {}
                # Try the most common FK name; if the field doesn't exist we
                # catch AttributeError and fall through to the zero default.
                try:
                    FoodOrder._meta.get_field("prepared_by")
                    food_qs_kwargs = dict(
                        prepared_by_id__in=kitchen_ids,
                        order_status="completed",
                        created_at__date__gte=start,
                        created_at__date__lte=end,
                    )
                    food_rows = (
                        FoodOrder.objects.filter(**food_qs_kwargs)
                        .values("prepared_by_id")
                        .annotate(
                            completed=Count("id"),
                            avg_prep=Avg(
                                F("completed_at") - F("created_at"),
                                output_field=models.DurationField(),
                            ),
                        )
                    )
                    for row in food_rows:
                        pid = row["prepared_by_id"]
                        if pid in perf:
                            perf[pid]["orders_completed"]    = row["completed"]
                            if row["avg_prep"] is not None:
                                perf[pid]["avg_preparation_time"] = round(
                                    row["avg_prep"].total_seconds() / 60, 1
                                )
                except Exception:
                    pass   # FoodOrder has no prepared_by FK — leave zeros

                # pending / cancelled orders (global counts for the period,
                # not per-staff since there's no staff FK yet)
                try:
                    kitchen_pending   = FoodOrder.objects.filter(
                        order_status="pending",
                        created_at__date__gte=start,
                        created_at__date__lte=end,
                    ).count()
                    kitchen_cancelled = FoodOrder.objects.filter(
                        order_status="cancelled",
                        created_at__date__gte=start,
                        created_at__date__lte=end,
                    ).count()
                    for pid in kitchen_ids:
                        if pid in perf:
                            perf[pid]["pending_orders"]   = kitchen_pending
                            perf[pid]["cancelled_orders"] = kitchen_cancelled
                except Exception:
                    pass
        except ImportError:
            pass   # food app not installed

        # ── 6. Security — IncidentLog metrics ────────────────────────────────────
        security_ids = [
            pid for pid, d in perf.items()
            if d["role"] == "security"
        ]
        if security_ids:
            from staff.models import IncidentLog

            # Total incidents logged per security staff member
            incident_counts = (
                IncidentLog.objects.filter(
                    logged_by_id__in=security_ids,
                    created_at__date__gte=start,
                    created_at__date__lte=end,
                )
                .values("logged_by_id")
                .annotate(
                    total=Count("id"),
                    resolved=Count("id", filter=Q(status="resolved")),
                    high=Count(
                        "id",
                        filter=Q(severity__in=["high", "critical"]),
                    ),
                )
            )
            for row in incident_counts:
                pid = row["logged_by_id"]
                if pid in perf:
                    perf[pid]["incidents_logged"]   = row["total"]
                    perf[pid]["incidents_resolved"] = row["resolved"]
                    perf[pid]["high_severity"]       = row["high"]

            # Average resolution time (resolved_at - created_at) in minutes
            resolution_times = (
                IncidentLog.objects.filter(
                    logged_by_id__in=security_ids,
                    created_at__date__gte=start,
                    created_at__date__lte=end,
                    status="resolved",
                    resolved_at__isnull=False,
                )
                .values("logged_by_id")
                .annotate(
                    avg_res=Avg(
                        F("resolved_at") - F("created_at"),
                        output_field=models.DurationField(),
                    )
                )
            )
            for row in resolution_times:
                pid = row["logged_by_id"]
                if pid in perf and row["avg_res"] is not None:
                    perf[pid]["avg_resolution_time"] = round(
                        row["avg_res"].total_seconds() / 60, 1
                    )

        # ── 7. Assemble rows and summary ──────────────────────────────────────
        rows = sorted(
            perf.values(),
            key=lambda r: -(
                r["check_ins_handled"]
                + r["rooms_cleaned"]
                + r["orders_completed"]
            ),
        )

        summary_full = {
            "total_staff":            len(rows),
            "total_check_ins":        sum(r["check_ins_handled"] for r in rows),
            "total_cleaning_done":    sum(r["rooms_cleaned"]      for r in rows),
            "total_maintenance_done": sum(
                r["orders_completed"] for r in rows
                if r["role"] == "maintenance"
            ),
            "total_incidents_logged": sum(
                r["incidents_logged"] for r in rows
                if r["role"] == "security"
            ),
            # Period boundaries for the summary band in StaffResultPanel
            "period_start": str(start),
            "period_end":   str(end),
        }

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }

    # ── Payments ───────────────────────────────────────────────────────────────

    @staticmethod
    def _payments(
        start: date, end: date,
        metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        """
        Queries payments.Payment (the dedicated payments table) — NOT Booking.

        Financial rules enforced here and visible via PaymentsResultPanel:
          - Only PAID payments count toward gross / net revenue.
          - FAILED payments are counted separately and never added to revenue.
          - REFUNDED payments reduce net_collected_amount via the Refund table.
          - PENDING payments are shown in their own card (awaiting completion).

        group_by options: day | week | month | payment_method | payment_status
        """
        from payments.models import Payment, PaymentStatus, Refund
        from django.db.models import DecimalField
        from django.db.models.functions import Coalesce

        qs = Payment.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        )

        # Optional filters
        if filters.get("room_type"):
            qs = qs.filter(booking__room__room_type=filters["room_type"])
        if filters.get("payment_status"):
            qs = qs.filter(status=filters["payment_status"])
        if filters.get("payment_method"):
            qs = qs.filter(payment_method=filters["payment_method"])

        # ── Aggregate KPIs ────────────────────────────────────────────────────
        paid_qs   = qs.filter(status=PaymentStatus.PAID)
        failed_qs = qs.filter(status=PaymentStatus.FAILED)
        pending_qs = qs.filter(status=PaymentStatus.PENDING)

        paid_agg = paid_qs.aggregate(
            gross=Coalesce(Sum("amount"), Decimal("0"), output_field=DecimalField()),
            count=Count("id"),
            avg=Coalesce(Avg("amount"), Decimal("0"), output_field=DecimalField()),
        )

        # Total refunds processed in the period (from the Refund table)
        refund_total = (
            Refund.objects.filter(
                status="completed",
                created_at__date__gte=start,
                created_at__date__lte=end,
            ).aggregate(
                t=Coalesce(Sum("amount"), Decimal("0"), output_field=DecimalField())
            )["t"]
        )

        gross_amount = paid_agg["gross"]
        net_amount   = gross_amount - refund_total
        total_txns   = qs.count()

        summary_full = {
            "total_gross_amount":       float(gross_amount),
            "net_collected_amount":     float(net_amount),
            "total_refunds":            float(refund_total),
            "successful_payments":      paid_agg["count"],
            "failed_payments":          failed_qs.count(),
            "pending_payments":         pending_qs.count(),
            "total_payments_processed": total_txns,
            "average_transaction_value": float(paid_agg["avg"]),
            "refund_rate": (
                round(float(refund_total) / float(gross_amount) * 100, 2)
                if gross_amount else 0.0
            ),
            "failed_payment_rate": (
                round(failed_qs.count() / total_txns * 100, 2)
                if total_txns else 0.0
            ),
        }

        # ── Rows (grouped) ────────────────────────────────────────────────────
        rows: list[dict] = []

        if group_by in ("day", "week", "month"):
            trunc = _trunc_fn(group_by)
            rows_qs = (
                qs.annotate(period=trunc("created_at"))
                  .values("period")
                  .annotate(
                      total_gross_amount=Coalesce(
                          Sum("amount", filter=Q(status=PaymentStatus.PAID)),
                          Decimal("0"), output_field=DecimalField(),
                      ),
                      successful_payments=Count("id", filter=Q(status=PaymentStatus.PAID)),
                      failed_payments=Count("id", filter=Q(status=PaymentStatus.FAILED)),
                      pending_payments=Count("id", filter=Q(status=PaymentStatus.PENDING)),
                  )
                  .order_by("period")
            )
            rows = [
                {
                    "period":               str(r["period"]),
                    "total_gross_amount":   float(r["total_gross_amount"]),
                    "net_collected_amount": float(r["total_gross_amount"]),  # refunds not per-period
                    "successful_payments":  r["successful_payments"],
                    "failed_payments":      r["failed_payments"],
                    "pending_payments":     r["pending_payments"],
                }
                for r in rows_qs
            ]

        elif group_by == "payment_method":
            rows_qs = (
                qs.values("payment_method")
                  .annotate(
                      total_gross_amount=Coalesce(
                          Sum("amount", filter=Q(status=PaymentStatus.PAID)),
                          Decimal("0"), output_field=DecimalField(),
                      ),
                      successful_payments=Count("id", filter=Q(status=PaymentStatus.PAID)),
                      failed_payments=Count("id", filter=Q(status=PaymentStatus.FAILED)),
                  )
                  .order_by("-total_gross_amount")
            )
            rows = [
                {
                    "payment_method":       r["payment_method"],
                    "total_gross_amount":   float(r["total_gross_amount"]),
                    "successful_payments":  r["successful_payments"],
                    "failed_payments":      r["failed_payments"],
                }
                for r in rows_qs
            ]

        elif group_by == "payment_status":
            rows_qs = (
                qs.values("status")
                  .annotate(
                      count=Count("id"),
                      total_amount=Coalesce(
                          Sum("amount"), Decimal("0"), output_field=DecimalField(),
                      ),
                  )
                  .order_by("-count")
            )
            rows = [
                {
                    "payment_status": r["status"],
                    "count":          r["count"],
                    "total_amount":   float(r["total_amount"]),
                }
                for r in rows_qs
            ]

        return {
            "summary": _filter_summary(summary_full, metrics),
            "rows":    rows,
        }


    @staticmethod
    def _food(
            start, end,
            metrics: list, group_by: str, filters: dict, user=None
    ) -> dict:
        from food.models import FoodOrder
        from django.db.models import Sum, Count, Avg
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth

        qs = FoodOrder.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        ).select_related('food_item')

        if filters.get('category'):
            qs = qs.filter(food_item__category=filters['category'])
        if filters.get('order_status'):
            qs = qs.filter(order_status=filters['order_status'])
        if filters.get('payment_type'):
            qs = qs.filter(payment_type=filters['payment_type'])

        totals = qs.aggregate(
            total_orders=Count('id'),
            total_revenue=Sum('total_price'),
            avg_order_value=Avg('total_price'),
        )
        paid_revenue = (
                qs.filter(payment_status='paid')
                .aggregate(t=Sum('total_price'))['t'] or 0
        )

        summary_full = {
            'total_orders': totals['total_orders'] or 0,
            'total_revenue': float(totals['total_revenue'] or 0),
            'paid_revenue': float(paid_revenue),
            'avg_order_value': float(totals['avg_order_value'] or 0),
            'pending_orders': qs.filter(order_status='pending').count(),
            'completed_orders': qs.filter(order_status='completed').count(),
        }

        # Grouping
        if group_by in ('day', 'week', 'month'):
            trunc = {'day': TruncDate, 'week': TruncWeek, 'month': TruncMonth}[group_by]
            rows_qs = (
                qs.annotate(period=trunc('created_at'))
                .values('period')
                .annotate(orders=Count('id'), revenue=Sum('total_price'))
                .order_by('period')
            )
            rows = [
                {
                    'period': str(r['period']),
                    'orders': r['orders'],
                    'revenue': float(r['revenue'] or 0),
                }
                for r in rows_qs
            ]
        elif group_by == 'category':
            cats_qs = (
                qs.values('food_item__category')
                .annotate(orders=Count('id'), revenue=Sum('total_price'))
                .order_by('-revenue')
            )
            rows = [
                {
                    'category': r['food_item__category'],
                    'orders': r['orders'],
                    'revenue': float(r['revenue'] or 0),
                }
                for r in cats_qs
            ]
        else:
            rows = []

        return {
            'summary': _filter_summary(summary_full, metrics),
            'rows': rows,
        }


# ─── Export helpers ───────────────────────────────────────────────────────────

# Human-readable column labels for every field the reports produce.
# Keys are the raw dict keys from rows[] and summary{}.
COLUMN_LABELS = {
    # Identity / period
    "period":                   "Period",
    "staff_id":                 "Staff ID",
    "name":                     "Name",
    "email":                    "Email",
    "role":                     "Role",
    # Bookings
    "bookings":                 "Bookings",
    "count":                    "Count",
    "status":                   "Status",
    "room_type":                "Room Type",
    "label":                    "Label",
    "paid_revenue":             "Paid Revenue (₱)",
    # Revenue
    "revenue":                  "Revenue (₱)",
    "total_revenue":            "Total Revenue (₱)",
    "total_tax":                "Total Tax (₱)",
    "total_service_fee":        "Service Fee (₱)",
    "net_revenue":              "Net Revenue (₱)",
    "avg_booking_value":        "Avg Booking Value (₱)",
    "paid_bookings":            "Paid Bookings",
    "total_refunds":            "Total Refunds (₱)",
    # Occupancy
    "room_count":               "Room Count",
    "room_type_label":          "Room Type",
    "occupied_nights":          "Occupied Nights",
    "max_nights":               "Max Nights",
    "occupancy_rate":           "Occupancy Rate (%)",
    "total_rooms":              "Total Rooms",
    "total_days":               "Total Days",
    "total_room_nights":        "Total Room Nights",
    # Guests
    "spent":                    "Total Spent (₱)",
    # Front desk
    "check_ins_handled":        "Check-ins Handled",
    "check_outs_handled":       "Check-outs Handled",
    "bookings_created":         "Bookings Created",
    "cancellations_processed":  "Cancellations Handled",
    "avg_check_in_time":        "Avg Check-in Time (min)",
    # Housekeeping
    "rooms_cleaned":            "Rooms Cleaned",
    "avg_cleaning_time":        "Avg Cleaning Time (min)",
    "rooms_cleaned_per_shift":  "Rooms / Shift",
    "delayed_cleanings":        "Delayed Cleanings",
    # Maintenance / kitchen
    "orders_completed":         "Tasks / Orders Completed",
    "avg_preparation_time":     "Avg Prep Time (min)",
    "pending_orders":           "Pending",
    "cancelled_orders":         "Cancelled",
    # Security
    "incidents_logged":         "Incidents Logged",
    "incidents_resolved":       "Incidents Resolved",
    "high_severity":            "High / Critical Incidents",
    "avg_resolution_time":      "Avg Resolution Time (min)",
    # Payments
    "payment_method":           "Payment Method",
    "payment_status":           "Payment Status",
    "total_gross_amount":       "Gross Amount (₱)",
    "net_collected_amount":     "Net Collected (₱)",
    "total_amount":             "Total Amount (₱)",
    "successful_payments":      "Successful Payments",
    "failed_payments":          "Failed Payments",
    "pending_payments":         "Pending Payments",
    "total_payments_processed": "Total Transactions",
    "average_transaction_value":"Avg Transaction (₱)",
    "refund_rate":              "Refund Rate (%)",
    "failed_payment_rate":      "Failure Rate (%)",
    # Food
    "category":                 "Category",
    "orders":                   "Orders",
    "total_orders":             "Total Orders",
    "avg_order_value":          "Avg Order Value (₱)",
    "paid_revenue":             "Paid Revenue (₱)",
    "pending_orders":           "Pending Orders",
    "completed_orders":         "Completed Orders",
    # Summary keys
    "total_check_ins":          "Total Check-ins",
    "total_cleaning_done":      "Total Rooms Cleaned",
    "total_maintenance_done":   "Total Maintenance Tasks",
    "total_incidents_logged":   "Total Incidents Logged",
    "total_staff":              "Total Staff",
    "period_start":             "Period Start",
    "period_end":               "Period End",
    "total_gross_amount":       "Gross Amount (₱)",
    "net_collected_amount":     "Net Collected (₱)",
}

# Fields that contain monetary amounts — formatted as ₱ in CSV/PDF,
# stored as real numbers in Excel so it can sum/chart them.
CURRENCY_FIELDS = {
    "revenue", "paid_revenue", "total_revenue", "total_tax",
    "total_service_fee", "net_revenue", "avg_booking_value",
    "total_refunds", "spent", "total_gross_amount",
    "net_collected_amount", "total_amount", "average_transaction_value",
    "avg_order_value", "paid_revenue",
}

# Fields that are percentages
PERCENT_FIELDS = {"occupancy_rate", "refund_rate", "failed_payment_rate"}

# Role labels for staff report grouping headers
ROLE_LABELS = {
    "front_desk":   "Front Desk",
    "housekeeping": "Housekeeping",
    "maintenance":  "Maintenance",
    "kitchen_staff":"Kitchen Staff",
    "security":     "Security",
}


def _col_label(key: str) -> str:
    """Return a human-readable column header for a field key."""
    return COLUMN_LABELS.get(key, key.replace("_", " ").title())


def _fmt_value(key: str, val) -> str:
    """Format a value for CSV / PDF text cells."""
    if val is None:
        return "—"
    if key in CURRENCY_FIELDS and isinstance(val, (int, float)):
        return f"₱{val:,.2f}"
    if key in PERCENT_FIELDS and isinstance(val, (int, float)):
        return f"{val:.1f}%"
    if isinstance(val, float):
        return f"{val:,.2f}" if val != int(val) else f"{int(val):,}"
    if isinstance(val, int):
        return f"{val:,}"
    return str(val)


def _is_staff_report(data: dict) -> bool:
    return data.get("meta", {}).get("report_type") == "staff"


def _group_staff_rows(rows: list) -> list[tuple[str, list]]:
    """
    Returns [(role_label, [rows]), ...] preserving ROLE_LABELS order.
    Used by CSV and PDF to write role-separated sections.
    """
    grouped: dict[str, list] = {}
    for row in rows:
        role = row.get("role", "unknown")
        grouped.setdefault(role, []).append(row)

    ordered = []
    for role_key in ROLE_LABELS:
        if role_key in grouped:
            ordered.append((ROLE_LABELS[role_key], grouped[role_key]))
    # Append any unknown/unexpected roles at the end
    for role_key, role_rows in grouped.items():
        if role_key not in ROLE_LABELS:
            ordered.append((role_key.replace("_", " ").title(), role_rows))
    return ordered


# ── Role-specific columns so each section only shows relevant columns ─────────

ROLE_COLUMNS = {
    "front_desk": [
        "staff_id", "name", "email",
        "check_ins_handled", "check_outs_handled",
        "bookings_created", "cancellations_processed", "avg_check_in_time",
    ],
    "housekeeping": [
        "staff_id", "name", "email",
        "rooms_cleaned", "avg_cleaning_time",
        "rooms_cleaned_per_shift", "delayed_cleanings",
    ],
    "maintenance": [
        "staff_id", "name", "email",
        "orders_completed", "pending_orders", "cancelled_orders",
    ],
    "kitchen_staff": [
        "staff_id", "name", "email",
        "orders_completed", "avg_preparation_time",
        "pending_orders", "cancelled_orders",
    ],
    "security": [
        "staff_id", "name", "email",
        "incidents_logged", "incidents_resolved",
        "high_severity", "avg_resolution_time",
    ],
}


def export_csv(data: dict, report_type: str) -> bytes:
    """
    Render report data as CSV bytes.

    Structure:
      Section 1 — Report info (type, period, generated_at)
      Section 2 — Summary (metric, value pairs)
      Section 3 — Data rows
        For staff reports: one sub-section per role with role-specific columns.
        For all others: single flat table.
    """
    import csv as _csv

    buf = io.StringIO()
    w   = _csv.writer(buf)

    meta = data.get("meta", {})

    # ── Header block ──────────────────────────────────────────────────────────
    w.writerow(["Report Type", report_type.replace("_", " ").title()])
    if meta.get("start_date"):
        w.writerow(["Period", f"{meta['start_date']} to {meta['end_date']}"])
    if meta.get("generated_at"):
        w.writerow(["Generated", meta["generated_at"]])
    w.writerow([])

    # ── Summary section ───────────────────────────────────────────────────────
    summary = data.get("summary", {})
    if summary:
        w.writerow(["SUMMARY"])
        w.writerow(["Metric", "Value"])
        for k, v in summary.items():
            w.writerow([_col_label(k), _fmt_value(k, v)])
        w.writerow([])

    # ── Data rows ─────────────────────────────────────────────────────────────
    rows = data.get("rows", [])
    if not rows:
        w.writerow(["No data available for the selected period and filters."])
        return buf.getvalue().encode("utf-8")

    if _is_staff_report(data):
        # Per-role sections with role-specific columns
        for role_label, role_rows in _group_staff_rows(rows):
            if not role_rows:
                continue
            role_key = role_rows[0].get("role", "")
            cols     = ROLE_COLUMNS.get(role_key, list(role_rows[0].keys()))
            # Only keep cols that actually exist in the rows
            cols = [c for c in cols if c in role_rows[0]]

            w.writerow([])
            w.writerow([f"── {role_label} ──"])
            w.writerow([_col_label(c) for c in cols])
            for row in role_rows:
                w.writerow([_fmt_value(c, row.get(c)) for c in cols])
    else:
        cols = list(rows[0].keys())
        w.writerow([_col_label(c) for c in cols])
        for row in rows:
            w.writerow([_fmt_value(c, row.get(c)) for c in cols])

    return buf.getvalue().encode("utf-8")


def export_pdf(data: dict, report_type: str) -> bytes:
    """
    Render report data as a PDF using ReportLab.
    Falls back to a plain-text document if ReportLab is unavailable.

    Layout:
      - Title + period line
      - Summary table (Metric | Value)
      - Data table(s) — one per role for staff reports
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units     import cm
        from reportlab.lib.enums     import TA_LEFT
        from reportlab.platypus      import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib import colors

        buf    = io.BytesIO()
        doc    = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm,  bottomMargin=1.5*cm,
        )
        styles  = getSampleStyleSheet()
        story   = []
        meta    = data.get("meta", {})
        summary = data.get("summary", {})
        rows    = data.get("rows", [])

        # Shared table style builder
        HEADER_BG  = colors.HexColor("#1a2744")
        ALT_ROW    = colors.HexColor("#f3f4f6")
        ROLE_COLORS = {
            "front_desk":   colors.HexColor("#1D4ED8"),
            "housekeeping": colors.HexColor("#065F46"),
            "maintenance":  colors.HexColor("#5B21B6"),
            "kitchen_staff":colors.HexColor("#92400E"),
            "security":     colors.HexColor("#991B1B"),
        }

        def _make_table(table_data, col_widths=None, header_color=HEADER_BG):
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), header_color),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 8),
                ("FONTSIZE",      (0, 1), (-1, -1), 7),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, ALT_ROW]),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING",   (0, 0), (-1, -1), 5),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ]))
            return t

        # ── Title ─────────────────────────────────────────────────────────────
        title_text = f"{report_type.replace('_', ' ').title()} Report"
        story.append(Paragraph(title_text, styles["Title"]))

        period_text = ""
        if meta.get("start_date"):
            period_text = (
                f"Period: {meta['start_date']} — {meta['end_date']}   |   "
                f"Generated: {meta.get('generated_at', '')[:19].replace('T', ' ')}"
            )
        if meta.get("group_by"):
            period_text += f"   |   Grouped by: {meta['group_by'].replace('_', ' ').title()}"
        if period_text:
            story.append(Paragraph(period_text, styles["Normal"]))
        story.append(Spacer(1, 0.4 * cm))

        # ── Summary table ─────────────────────────────────────────────────────
        if summary:
            story.append(Paragraph("Summary", styles["Heading2"]))
            s_data = [["Metric", "Value"]] + [
                [_col_label(k), _fmt_value(k, v)]
                for k, v in summary.items()
                if k not in ("period_start", "period_end")  # shown in header already
            ]
            story.append(_make_table(s_data, col_widths=[9*cm, 7*cm]))
            story.append(Spacer(1, 0.5 * cm))

        # ── Data table(s) ─────────────────────────────────────────────────────
        if not rows:
            story.append(Paragraph(
                "No data available for the selected period and filters.",
                styles["Normal"],
            ))
        elif _is_staff_report(data):
            story.append(Paragraph("Staff Performance by Role", styles["Heading2"]))
            page_width = landscape(A4)[0] - 3*cm  # usable width

            for role_label, role_rows in _group_staff_rows(rows):
                if not role_rows:
                    continue
                role_key     = role_rows[0].get("role", "")
                header_color = ROLE_COLORS.get(role_key, HEADER_BG)
                cols         = ROLE_COLUMNS.get(role_key, list(role_rows[0].keys()))
                cols         = [c for c in cols if c in role_rows[0]]

                story.append(Spacer(1, 0.3 * cm))
                story.append(Paragraph(role_label, styles["Heading3"]))
                story.append(Spacer(1, 0.15 * cm))

                col_w = page_width / len(cols)
                t_data = [[_col_label(c) for c in cols]] + [
                    [_fmt_value(c, row.get(c)) for c in cols]
                    for row in role_rows
                ]
                story.append(_make_table(t_data, col_widths=[col_w]*len(cols),
                                         header_color=header_color))
        else:
            story.append(Paragraph("Data", styles["Heading2"]))
            cols      = list(rows[0].keys())
            page_width = landscape(A4)[0] - 3*cm
            col_w     = max(2*cm, page_width / len(cols))
            t_data    = [[_col_label(c) for c in cols]] + [
                [_fmt_value(c, row.get(c)) for c in cols]
                for row in rows
            ]
            story.append(_make_table(t_data, col_widths=[col_w]*len(cols)))

        doc.build(story)
        return buf.getvalue()

    except ImportError:
        # Fallback: structured plain text
        lines = [
            f"REPORT: {report_type.replace('_', ' ').upper()}",
            f"Period: {data.get('meta', {}).get('start_date', '')} to "
            f"{data.get('meta', {}).get('end_date', '')}",
            "",
            "SUMMARY",
            "--------",
        ]
        for k, v in data.get("summary", {}).items():
            lines.append(f"{_col_label(k)}: {_fmt_value(k, v)}")

        lines += ["", "DATA", "----"]
        rows = data.get("rows", [])
        if rows:
            if _is_staff_report(data):
                for role_label, role_rows in _group_staff_rows(rows):
                    lines.append(f"\n[{role_label}]")
                    cols = ROLE_COLUMNS.get(
                        role_rows[0].get("role", ""), list(role_rows[0].keys())
                    )
                    cols = [c for c in cols if c in role_rows[0]]
                    lines.append("\t".join(_col_label(c) for c in cols))
                    for row in role_rows:
                        lines.append("\t".join(_fmt_value(c, row.get(c)) for c in cols))
            else:
                cols = list(rows[0].keys())
                lines.append("\t".join(_col_label(c) for c in cols))
                for row in rows:
                    lines.append("\t".join(_fmt_value(c, row.get(c)) for c in cols))
        else:
            lines.append("No data.")

        return "\n".join(lines).encode("utf-8")


def export_excel(data: dict, report_type: str) -> bytes:
    """
    Render report data as an Excel workbook using openpyxl.

    Sheets:
      Summary  — KPI summary with human-readable labels and number formatting
      Data     — Flat rows for non-staff reports; one sheet per role for staff
      Meta     — Report metadata (type, period, generated_at, filters)

    Improvements over original:
      - Human-readable column headers via COLUMN_LABELS
      - Currency fields stored as real numbers with ₱ accounting format
      - Percentage fields stored as real numbers with % format
      - Integer fields stored as integers
      - Freeze panes on row 1 of every data sheet
      - Auto column widths (capped at 40)
      - Staff reports split into per-role sheets with role-specific columns
      - Non-data rows (empty values) still present as 0, not ""
    """
    try:
        import openpyxl
        from openpyxl.styles  import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils   import get_column_letter

        HEADER_FONT  = Font(bold=True, color="FFFFFF", size=9)
        HEADER_FILL  = PatternFill("solid", fgColor="1A2744")
        ALT_FILL     = PatternFill("solid", fgColor="F3F4F6")
        ROLE_FILLS   = {
            "front_desk":    PatternFill("solid", fgColor="1D4ED8"),
            "housekeeping":  PatternFill("solid", fgColor="065F46"),
            "maintenance":   PatternFill("solid", fgColor="5B21B6"),
            "kitchen_staff": PatternFill("solid", fgColor="92400E"),
            "security":      PatternFill("solid", fgColor="991B1B"),
        }
        THIN_BORDER  = Border(
            bottom=Side(style="thin", color="D1D5DB"),
        )

        PHP_FMT  = '₱#,##0.00'
        PCT_FMT  = '0.0"%"'
        INT_FMT  = '#,##0'
        FLOAT_FMT= '#,##0.00'

        def _cell_value(key, val):
            """Return (native_value, number_format) for a cell."""
            if val is None:
                return (0 if key not in ("name", "email", "role", "staff_id",
                                         "period", "payment_method",
                                         "payment_status", "room_type",
                                         "category", "status") else "—",
                        None)
            if key in CURRENCY_FIELDS and isinstance(val, (int, float)):
                return (float(val), PHP_FMT)
            if key in PERCENT_FIELDS and isinstance(val, (int, float)):
                return (float(val), PCT_FMT)
            if isinstance(val, float):
                return (val, FLOAT_FMT)
            if isinstance(val, int):
                return (val, INT_FMT)
            return (val, None)

        def _write_sheet(ws, cols, rows, header_fill=HEADER_FILL):
            """Write a header row + data rows to a worksheet."""
            # Header
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=_col_label(col))
                cell.font      = HEADER_FONT
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows
            for ri, row in enumerate(rows, 2):
                fill = ALT_FILL if ri % 2 == 0 else None
                for ci, col in enumerate(cols, 1):
                    val, fmt = _cell_value(col, row.get(col))
                    cell      = ws.cell(row=ri, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if fmt:
                        cell.number_format = fmt
                    if fill:
                        cell.fill = fill

            # Freeze header row
            ws.freeze_panes = ws.cell(row=2, column=1)

            # Auto column widths
            for ci, col in enumerate(cols, 1):
                letter   = get_column_letter(ci)
                max_len  = len(_col_label(col))
                for ri in range(2, min(len(rows) + 2, 500)):  # sample first 500
                    cell_val = ws.cell(row=ri, column=ci).value
                    if cell_val is not None:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[letter].width = min(max_len + 3, 40)

        wb      = openpyxl.Workbook()
        meta    = data.get("meta", {})
        summary = data.get("summary", {})
        rows    = data.get("rows", [])

        # ── Summary sheet ──────────────────────────────────────────────────────
        ws_summary       = wb.active
        ws_summary.title = "Summary"

        # Report info block at the top
        info_rows = [
            ("Report Type",  report_type.replace("_", " ").title()),
            ("Period",       f"{meta.get('start_date', '')} to {meta.get('end_date', '')}"),
            ("Group By",     meta.get("group_by", "").replace("_", " ").title()),
            ("Generated",    str(meta.get("generated_at", ""))[:19].replace("T", " ")),
        ]
        for r, (label, value) in enumerate(info_rows, 1):
            ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True, size=9)
            ws_summary.cell(row=r, column=2, value=value)

        # Blank row
        info_end = len(info_rows) + 2

        # Summary header
        hdr_row = info_end
        ws_summary.cell(row=hdr_row, column=1, value="Metric").font   = HEADER_FONT
        ws_summary.cell(row=hdr_row, column=1).fill  = HEADER_FILL
        ws_summary.cell(row=hdr_row, column=2, value="Value").font    = HEADER_FONT
        ws_summary.cell(row=hdr_row, column=2).fill  = HEADER_FILL

        for si, (k, v) in enumerate(summary.items(), hdr_row + 1):
            ws_summary.cell(row=si, column=1, value=_col_label(k))
            val, fmt = _cell_value(k, v)
            cell     = ws_summary.cell(row=si, column=2, value=val)
            if fmt:
                cell.number_format = fmt
            if si % 2 == 0:
                ws_summary.cell(row=si, column=1).fill = ALT_FILL
                cell.fill = ALT_FILL

        ws_summary.column_dimensions["A"].width = 32
        ws_summary.column_dimensions["B"].width = 22

        # ── Data sheet(s) ──────────────────────────────────────────────────────
        if rows:
            if _is_staff_report(data):
                for role_label, role_rows in _group_staff_rows(rows):
                    if not role_rows:
                        continue
                    role_key  = role_rows[0].get("role", "")
                    sheet_name = role_label[:31]   # Excel sheet name limit
                    ws         = wb.create_sheet(title=sheet_name)
                    cols       = ROLE_COLUMNS.get(role_key, list(role_rows[0].keys()))
                    cols       = [c for c in cols if c in role_rows[0]]
                    h_fill     = ROLE_FILLS.get(role_key, HEADER_FILL)
                    _write_sheet(ws, cols, role_rows, header_fill=h_fill)
            else:
                ws   = wb.create_sheet(title="Data")
                cols = list(rows[0].keys())
                _write_sheet(ws, cols, rows)

        # ── Meta sheet ─────────────────────────────────────────────────────────
        ws_meta       = wb.create_sheet(title="Meta")
        ws_meta.cell(row=1, column=1, value="Key").font  = HEADER_FONT
        ws_meta.cell(row=1, column=1).fill = HEADER_FILL
        ws_meta.cell(row=1, column=2, value="Value").font = HEADER_FONT
        ws_meta.cell(row=1, column=2).fill = HEADER_FILL
        for mi, (k, v) in enumerate(meta.items(), 2):
            ws_meta.cell(row=mi, column=1, value=k)
            ws_meta.cell(row=mi, column=2, value=str(v))
        ws_meta.column_dimensions["A"].width = 20
        ws_meta.column_dimensions["B"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except ImportError:
        return export_csv(data, report_type)


# ─── Execution runner (moved here from views.py to avoid circular imports) ────
# tasks.py imports this directly; views.py also imports and re-uses it.

def _notify_report_ready(user, report_type, execution, is_scheduled, export_format):
    """
    Creates a dashboard notification for the report owner when a report
    completes — used for scheduled runs and file-export (CSV/PDF/Excel) runs.
    """
    try:
        from notifications.models import (
            Notification,
            NotificationEvent,
            NotificationChannel,
            NotificationRecipientType,
            NotificationPriority,
        )
        trigger   = "Scheduled report" if is_scheduled else "Your report"
        fmt_label = {"csv": "CSV", "pdf": "PDF", "excel": "Excel", "json": "In-App"}.get(
            export_format, export_format.upper()
        )
        type_label = report_type.replace("_", " ").title()
        Notification.objects.create(
            recipient      = user,
            booking        = None,
            event          = NotificationEvent.SYSTEM_ALERT,
            recipient_type = NotificationRecipientType.ADMIN,
            channel        = NotificationChannel.DASHBOARD,
            priority       = NotificationPriority.MEDIUM,
            title          = f"Report Ready — {type_label}",
            description    = (
                f"{trigger} '{type_label}' ({fmt_label}) has completed successfully. "
                f"Go to Reports → History to view or download the results."
            ),
        )
    except Exception as exc:
        logger.warning(
            "Failed to send report-ready notification for execution %s: %s",
            execution.pk, exc,
        )


def _notify_report_failed(user, report_type, execution):
    """Creates a HIGH priority notification when a scheduled report fails."""
    try:
        from notifications.models import (
            Notification,
            NotificationEvent,
            NotificationChannel,
            NotificationRecipientType,
            NotificationPriority,
        )
        type_label = report_type.replace("_", " ").title()
        Notification.objects.create(
            recipient      = user,
            booking        = None,
            event          = NotificationEvent.SYSTEM_ALERT,
            recipient_type = NotificationRecipientType.ADMIN,
            channel        = NotificationChannel.DASHBOARD,
            priority       = NotificationPriority.HIGH,
            title          = f"Report Failed — {type_label}",
            description    = (
                f"Scheduled '{type_label}' report (execution #{execution.pk}) failed. "
                f"Error: {execution.error_message or 'Unknown error'}. "
                f"Check Reports → History for details."
            ),
        )
    except Exception as exc:
        logger.warning(
            "Failed to send report-failed notification for execution %s: %s",
            execution.pk, exc,
        )


def run_report_and_log(
    report_type: str,
    config: dict,
    export_format: str,
    user,
    template=None,
    schedule=None,
    is_scheduled: bool = False,
):
    """
    Execute a report, persist a ReportExecution log, write a StaffActivityLog
    audit entry, and fire a notification on completion or failure.

    Returns (data, execution).
    Imported by both views.py and tasks.py — keeping it here prevents the
    circular import that occurred when tasks.py imported from views.py.
    """
    from .models import ReportExecution, ExecutionStatus, ExportFormat

    execution = ReportExecution.objects.create(
        template        = template,
        schedule        = schedule,
        report_type     = report_type,
        config_snapshot = config,
        export_format   = export_format,
        triggered_by    = user,
        is_scheduled    = is_scheduled,
        status          = ExecutionStatus.PENDING,
    )

    # ── Audit log ─────────────────────────────────────────────────────────────
    try:
        from staff.models import StaffActivityLog
        profile = getattr(user, "staff_profile", None) if user else None
        period  = config.get("period", "—")
        filters = config.get("filters", {})
        metrics = config.get("metrics", [])

        StaffActivityLog.objects.create(
            staff       = profile,
            action_type = "generate_report",
            description = (
                f"Generated '{report_type}' report. "
                f"Period: {period}. Format: {export_format}. "
                f"Metrics: {metrics or 'all'}. Filters: {filters or 'none'}. "
                f"{'Scheduled run.' if is_scheduled else 'On-demand run.'}"
            ),
            metadata = {
                "report_type":   report_type,
                "period":        period,
                "export_format": export_format,
                "metrics":       metrics,
                "filters":       filters,
                "execution_id":  execution.pk,
                "is_scheduled":  is_scheduled,
                "template_id":   template.pk if template else None,
            },
        )
    except Exception as exc:
        logger.warning("StaffActivityLog write failed for execution %s: %s", execution.pk, exc)

    # ── Run ───────────────────────────────────────────────────────────────────
    try:
        data = EnhancedReportService.run(report_type, config, user=user)
        execution.mark_success(data)

        if user and (is_scheduled or export_format != ExportFormat.JSON):
            _notify_report_ready(
                user          = user,
                report_type   = report_type,
                execution     = execution,
                is_scheduled  = is_scheduled,
                export_format = export_format,
            )

        return data, execution

    except Exception as exc:
        logger.exception("Report execution %s failed: %s", execution.pk, exc)
        execution.mark_failed(str(exc))

        if user and is_scheduled:
            _notify_report_failed(
                user        = user,
                report_type = report_type,
                execution   = execution,
            )

        return None, execution